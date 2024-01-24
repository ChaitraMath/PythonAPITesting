import openpyxl
import requests
import json
import jsonpath

def test_add_multiple_students():
    api_url = "https://thetestingworldapi.com/api/studentsDetails"
    f = open("C:/Users/Chaitra/PycharmProjects/APIAutomation/addNewSt.json",'r')
    json_requests = json.loads(f.read())

    workbook = openpyxl.load_workbook("C:/Users/Chaitra/Desktop/pytest/TestData.xlsx")
    sheetname= workbook['Sheet1']
    rows = sheetname.max_row
    for i in range(2,rows+1):
        cell_first_name = sheetname.cell(row=i, column=1)
        cell_mid_name = sheetname.cell(row=i, column=2)
        cell_last_name = sheetname.cell(row=i, column=3)
        cell_dob = sheetname.cell(row=i, column=4)

        json_requests['first_name'] = cell_first_name.value
        json_requests['middle_name'] = cell_mid_name.value
        json_requests['last_name'] = cell_last_name.value
        json_requests['date_of_birth'] = cell_dob.value

        response = requests.post(api_url,json_requests)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201