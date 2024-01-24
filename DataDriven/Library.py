import json
import requests
import openpyxl
import jsonpath

class Common:

    def __init__(self,FileNamePath, SheetName):
        global wk
        global sh
        wk = openpyxl.load_workbook(FileNamePath)
        sh = wk[SheetName]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        col = sh.max_column
        return col

    def fetch_key_names(self):
        c = sh.max_column
        llist = []
        for i in range(1,c+1):
            cell_name = sh.cell(row=1,column=i)
            llist.insert(i-1, cell_name.value)
        return llist

    def update_request_with_data(self,rowNumber, jsonRequest, keyList):
        c = sh.max_column
        for i in range(1,c+1):
            cell_name = sh.cell(row=rowNumber,column=i)
            jsonRequest[keyList[i-1]] = cell_name.value
        return jsonRequest
